import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill  # 엑셀 스타일


class ClassificationExcel:
    # Constructor
    def __init__(self, order_xlsx_filename, partner_info_xlsx_filename):
        self.path = './data/'
        self.set_order_list(order_xlsx_filename)
        df_partners_info = pd.read_excel(partner_info_xlsx_filename, engine='openpyxl')

        # 브랜드에 값이 비어있다면 제거
        df_partners_info.dropna(axis='index', how='any', subset=['브랜드'])

        # 브랜드 list 생성
        self.brands = df_partners_info['브랜드'].tolist()

        # 업체명 list 생성
        self.partners = df_partners_info['업체명'].tolist()


    def set_order_list(self, order_xlsx_filename):
        # sendRequest 불러오기
        df = pd.read_excel(order_xlsx_filename, engine='openpyxl')

        # df파일 1번째 행을 칼럼으로 지정
        df = df.rename(columns=df.iloc[1])

        # 0~1행 삭제
        processed_df = df.drop([df.index[0], df.index[1]])

        # 인덱스 재설정
        self.order_list = processed_df.reset_index(drop=True)  # 가공한 df파일, #find_product에서 활용

    # 각 상품명에 brands 요소가 있는지 모두 확인하여 partners 이름으로 각각 excel 저장 기능
    def classify(self):
        print(self.brands)

        now_today = datetime.now().strftime('%Y-%m-%d')  # 일자

        for _, row in self.order_list.iterrows():  # 엑셀 주문 건수 갯수만큼 돌리기
            brand_name = ''
            partner_name = ''
            for j in range(len(self.brands)):  # brands의 갯수만큼 돌리기
                if self.brands[j] in row['상품명']:  # 전체 df 상품명에서 brands 값명이 있으면, 더클래스가 '차바치 더클래스'가 있는 주문을 제외하고
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break

            if brand_name != '':
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                # print(df_filtered)  # 여기까지는 모든 data가 나옴
                if len(df_filtered) != 0:  # df가 비어 있지 않으면
                    df_filtered.to_excel(f'{self.path}[패스트몰]발주요청서_{now_today}_{partner_name}.xlsx',
                                      index=False)  # partners 이름으로 excel 저장, index없이 저장
            else:
                print(row['상품명'])
                print('브랜드 이름을 찾을 수 없습니다.')
    # 저장한 주문파일 엑셀 폼 설정
    def set_excel_form(self):
        now_today = datetime.now().strftime('%Y-%m-%d')  # 일자
        # 폴더 안의 엑셀 파일 하나씩 불러오기

        file_list = os.listdir(self.path)  # path폴더에 있는 파일을 리스트로 받기
        print(file_list)

        for file_name_raw in file_list:
            file_name = f'{self.path}' + file_name_raw
            wb = load_workbook(filename=file_name)  # 엑셀파일 가져오기
            ws = wb.active  # 활성화

            # A1내용 삽입
            m_row = ws.max_row - 1  # 주문 건수 세기(칼럼행 제외)
            print(m_row)
            # 1,2행을 새로 삽입한다
            ws.insert_rows(1)
            ws.insert_rows(2)
            ws['A1'].value = f'발송요청내역 [총 {m_row}건, 1페이지] {now_today}'  # 전체 행 개수를 세서 num 변수에 삽입
            f = Font(size=11, bold=True)  # 1행 폰트 크기 11, bold
            ws['A1'].font = f
            # # A:U열 설정
            ws.merge_cells('A1:U1')  # 셀병합
            ws['A1'].alignment = Alignment(horizontal='left')  # 왼쪽 정렬
            f2 = Font(size=9)

            # 3, 4행 설정
            for row in ws.rows:
                for cell in row:
                    if cell.row != 1 and cell.row != 2:  # 1,2행 제외
                        cell.alignment = Alignment(horizontal='center', vertical='center',
                                                   wrap_text=True)  # 셀의 너비에 맞게 자동 줄바꿈, 모두 가운데 정렬
                        cell.fill = PatternFill(fgColor='ffffcc', fill_type='solid')  # 노란색 채우기
                        cell.font = Font(size=9)  # 3행부터 폰트크기 9
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                             bottom=Side(style='thin'))
                        if cell.row == 3:  # 3행이면
                            cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True,
                                                       wrap_text=False)  # 줄바꿈 제외, 셀 크기에 맞게 글자 축소
                            cell.font = Font(bold=True)  # 볼드체
                            cell.fill = PatternFill(fgColor='c0c0c0', fill_type='solid')  # 회색 채우기

            # 열 너비 설정
            columns_15 = [1, 21]
            columns_35 = [9, 10, 20]
            columns_10 = [2, 3, 4, 5, 6, 7, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19]
            for col in range(len(columns_15)):
                ws.column_dimensions[get_column_letter(columns_15[col])].width = 17  # A, U, 15로 하면 14.38이 됨
            for col in range(len(columns_35)):
                ws.column_dimensions[get_column_letter(columns_35[col])].width = 38
            for col in range(len(columns_10)):
                ws.column_dimensions[get_column_letter(columns_10[col])].width = 12
                # wb.save(f'{self.path}[패스트몰]발주요청서_2022-02-24_AUTOCOS.xlsx')#예시
            wb.save(f'{self.path}' + file_name_raw)
            print('엑셀폼 변경 완료')


# 인스턴스 생성
CE = ClassificationExcel('주문목록20221112.xlsx', '파트너목록.xlsx')

# #메소드 호출
CE.classify()
CE.set_excel_form()
