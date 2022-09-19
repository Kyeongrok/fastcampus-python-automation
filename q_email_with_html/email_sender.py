import os
import smtplib
from email.encoders import encode_base64
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from os.path import basename
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from datetime import datetime


class EmailSender:
    email_addr = None
    manager_name = None
    password = None
    smtp_server_map = {
        'gmail.com':'smtp.gmail.com',
        'naver.com':'smtp.naver.com'
    }
    smtp_server = None
    template_filename = None

    def __init__(self, email_addr, password, manager_name, template_filename, attachment_path='data/'):
        print('생성자')
        self.email_addr = email_addr
        self.manager_name = manager_name
        if password == None or '':
            raise Exception('password를 입력해주세요.')
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]] # fc.krkim@gmail.com
        print(self.smtp_server)
        self.template_filename = template_filename
        self.attachments_path = attachment_path

    def send_email(self, html_msg, from_addr, to_addr, receiver_name, subject, attachment):

        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEMultipart('alternative')
            msg['From'] = formataddr((self.manager_name, from_addr))
            msg['To'] = formataddr((receiver_name, to_addr))
            msg['Subject'] = subject + str(datetime.now())
            msg.attach(MIMEText(html_msg, 'html', 'utf-8'))
            if attachment:
                filenm = Path(attachment).name
                print('filenm:', filenm)
                print('basename:', basename(self.attachments_path + attachment))
                with open(self.attachments_path + attachment, 'rb') as f:
                    part = MIMEBase('application', "octet-steam")
                    part.set_payload(f.read())
                    # part.add_header('Content-Disposition', 'attachment; filename="%s"' % filenm)
                    part.add_header('content-disposition', 'attachment', filename='%s' % filenm)
                    encode_base64(part)
                    msg.attach(part)

            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print(f'to_addr:{to_addr}로 이메일 전송이 완료 되었습니다.')

    def send_all_emails(self, filename):
        print(f'{filename}에 있는 이메일과 내용을 이용해 메일을 보냅니다.')
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value != None:
                with open(self.template_filename, encoding='utf-8') as f:
                    temp1 = f.read()
                    print(row[0].value, row[1].value, row[2].value)
                    temp1 = temp1.replace('%받는분%', row[1].value)
                    temp1 = temp1.replace('%매니저_이름%', self.manager_name)
                    self.send_email(html_msg=temp1,
                                  from_addr=self.email_addr,
                                  to_addr=row[0].value, receiver_name=row[1].value,
                                  subject=row[2].value, attachment=row[3].value)
            else:
                print('row[0]이 None입니다.')

if __name__ == '__main__':
    # es = EmailSender('fc.krkim@gmail.com', os.getenv('MY_GMAIL_PASSWORD'))
    es = EmailSender('oceanfog@naver.com', os.getenv('MY_NAVER_PASSWORD'), manager_name='김미령')
    es.send_all_emails('이메일리스트_with_name.xlsx')