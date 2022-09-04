from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string

import pandas as pd
from datetime import datetime, timedelta
import calendar



class WeeklyWorkPlan:
    wb = None
    ws = None
    filename = '주간업무계획표.xlsx'
    dt_list = []
    days_of_week = []
    start_date = '2022-09-01'
    title = '주간업무계획표'
    manager = '홍길동'

    def __init__(self, filename, start_date, manager, sheet_no=0):
        self.filename = filename
        self.wb = Workbook()
        # 현재 활성화 되어있는 워크시트 기본값 1번째
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager
        self.set_dates()
        self.set_title()
        self.insert_context()
        self.style_setting()

    # workbook 만들기
    def create_excel_file(self, filename):
        self.wb.save(filename)
        print('엑셀파일 생성 완료')

    def save(self, filename):
        self.wb.save(filename)

    def set_dates(self):
        dt = datetime.strptime(self.start_date, '%Y-%m-%d') + timedelta(days=6)
        week = pd.date_range(start=datetime.strptime(self.start_date, '%Y-%m-%d'), end=dt.strftime("%Y%m%d"))
        dt_list = week.strftime("%Y-%m-%d").to_list()
        self.days_of_week = week.strftime("%A").to_list()
        self.dt_list = dt_list

    def set_title(self):

        ws = self.ws
        ws.title = self.title
        ws.cell(row=2, column=2, value='담당자')

        # ws.cell(2, 2, value='담당자')
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        ws['B5'] = self.title
        ws['B6'] = f'({self.dt_list[0]}~{self.dt_list[-1]})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')

    def insert_context(self):

        ws = self.ws

        cols_data = ['날짜', '요일', '시간', '일정', '비고']

        for col_idx in range(len(cols_data)):
            # print(col)
            ws.cell(row=8, column=2 + col_idx).value = cols_data[col_idx]


        row_num = 9
        s = 2
        for week in self.dt_list:
            ws.cell(row=row_num, column=s).value = week
            row_num = row_num + 1

        v = 9
        m = 3
        for day in self.days_of_week:
            ws.cell(row=v, column=m).value = day
            v = v + 1

        # 행 중간 삽입
        ws.insert_rows(10, 4)
        ws.insert_rows(15, 4)
        ws.insert_rows(20, 4)
        ws.insert_rows(25, 4)
        ws.insert_rows(30, 4)
        ws.insert_rows(35, 4)
        ws.insert_rows(40, 4)

        # 날짜, 요일, 비고 셀 병합

        ws.merge_cells('B9:B13')  # 날짜
        ws.merge_cells('C9:C13')  # 요일
        ws.merge_cells('F9:F13')  # 비고

        ws.merge_cells('B14:B18')  # 날짜
        ws.merge_cells('C14:C18')  # 요일
        ws.merge_cells('F14:F18')  # 비고

        ws.merge_cells('B19:B23')  # 날짜
        ws.merge_cells('C19:C23')  # 요일
        ws.merge_cells('F19:F23')  # 비고

        ws.merge_cells('B24:B28')  # 날짜
        ws.merge_cells('C24:C28')  # 요일
        ws.merge_cells('F24:F28')  # 비고

        ws.merge_cells('B29:B33')  # 날짜
        ws.merge_cells('C29:C33')  # 요일
        ws.merge_cells('F29:F33')  # 비고

        ws.merge_cells('B34:B38')  # 날짜
        ws.merge_cells('C34:C38')  # 요일
        ws.merge_cells('F34:F38')  # 비고

        ws.merge_cells('B39:B43')  # 날짜
        ws.merge_cells('C39:C43')  # 요일
        ws.merge_cells('F39:F43')  # 비고

        print('context 생성 완료')

    def style_setting(self):

        # 활성화
        ws = self.wb.worksheets[0]

        # B C D E F 열의 너비 15
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # E 일정 열너비 40 설정
        ws.column_dimensions['E'].width = 40

        # A 열너비 5 설정
        ws.column_dimensions['A'].width = 5

        # 주간업무계획표 폰트 크기 키우기, 굵게
        title_f = Font(name='맑은고딕', size=28, bold=True)
        # color='ff9999', strikethrough=True, underline='single', italic=True
        ws['B5'].font = title_f

        # 칼럼명 폰트 굵게
        font_range = ws['B8:F8']

        for col in font_range:
            for cell in col:
                cell.font = Font(bold=True)

        # 주간업무계획표, 기간 가운데 정렬
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

        # 칼럼명 가운데 정렬
        for col in font_range:
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 날짜, 요일 가운데 정렬
        for i in range(9, 40, 5):
            ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{i}'].alignment = Alignment(horizontal='center', vertical='center')

        # # 담당자, 시작일, 컬럼명 색칠
        ws['B2'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')
        ws['B3'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')

        # 표 컬럼명 색칠
        for col in ws['B8:F8']:
            for cell in col:
                if cell.row == 8:  # 8행이면
                    cell.fill = PatternFill(fgColor='E2EFDA', fill_type='solid')


        # 테두리 설정
        border_thin = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # 타이틀 영역
        for col in ws.iter_cols(min_row=2, min_col=2, max_row=3, max_col=3):
            for cell in col:
                cell.border = border_thin

        # context 영역
        for col in ws.iter_cols(min_row=8, min_col=2, max_row=43, max_col=6):
            for cell in col:
                cell.border = border_thin

        # 파일 저장하기
        print('스타일 적용 완료')


if __name__ == '__main__':
    file_name = '주간업무계획표.xlsx'
    wwp = WeeklyWorkPlan(file_name, start_date='2022-09-01', manager='김경록')
    wwp.save(file_name)
