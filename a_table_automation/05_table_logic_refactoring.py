from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string

import pandas as pd
from datetime import datetime, timedelta
import calendar


class WeeklyWorkPlan:
    wb = None
    ws = None
    dt_list = []
    days_of_week = []
    start_date = '2022-09-01'
    title = '주간업무계획표'
    manager = '홍길동'

    def __init__(self, start_date, manager, sheet_no=0):
        self.wb = Workbook()
        # 현재 활성화 되어있는 워크시트 기본값 1번째
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager
        self.set_dates()
        self.set_title()
        self.set_table()
        self.style_setting()

    def save(self, filename):
        self.wb.save(filename)
        print('엑셀파일 생성 완료')

    def set_dates(self):
        dt = datetime.strptime(self.start_date, '%Y-%m-%d') + timedelta(days=6)
        week = pd.date_range(start=datetime.strptime(self.start_date, '%Y-%m-%d'), end=dt.strftime("%Y%m%d"))
        dt_list = week.strftime("%Y-%m-%d").to_list()
        self.days_of_week = week.strftime("%A").to_list()
        self.dt_list = dt_list

        print(self.dt_list)
        print(self.days_of_week)

    def set_title(self):

        ws = self.ws
        ws.title = self.title
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        ws['B5'] = self.title
        # start ~ end날짜 설정
        ws['B6'] = f'({self.dt_list[0]} ~ {self.dt_list[-1]})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')

    def set_table(self):

        ws = self.ws

        # 컬럼명 채우기
        cols_data = ['날짜', '요일', '시간', '일정', '비고']
        for col_idx in range(len(cols_data)):
            ws.cell(row=8, column=2 + col_idx).value = cols_data[col_idx]

        # 날짜, 요일 채우기
        for i in range(len(self.dt_list)):
            row_idx = 9 + i * 5
            ws.cell(row=row_idx, column=2).value = self.dt_list[i]
            ws.cell(row=row_idx, column=3).value = self.days_of_week[i]
            # 날짜, 요일, 비고 셀 병합
            ws.merge_cells(f'B{row_idx}:B{row_idx + 4}')  # 날짜
            ws.merge_cells(f'C{row_idx}:C{row_idx + 4}')  # 요일
            ws.merge_cells(f'F{row_idx}:F{row_idx + 4}')  # 비고

        print('context 생성 완료')

    def style_setting(self):

        # 활성화
        ws = self.ws

        # B C D E F 열의 너비 15
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # E 일정 열너비 40 설정
        ws.column_dimensions['E'].width = 40

        # A 열너비 5 설정
        ws.column_dimensions['A'].width = 5

        # 주간업무계획표 폰트 크기 키우기, 굵게
        title_f = Font(name='맑은 고딕', size=28, bold=True)
        # color='ff9999', strikethrough=True, underline='single', italic=True
        ws['B5'].font = title_f

        for col in ws['B8:F8']:
            for cell in col:
                # 컬럼명 폰트 굵게
                cell.font = Font(bold=True)
                # 컬럼명 가운데 정렬
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 주간업무계획표, 기간 가운데 정렬
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')


        # 날짜, 요일 가운데 정렬
        for i in range(9, 40, 5):
            ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{i}'].alignment = Alignment(horizontal='center', vertical='center')

        # # 담당자, 시작일, 컬럼명 색칠
        ws['B2'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')
        ws['B3'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')

        # 표 컬럼명 색칠
        for col in ws['B8:F8']:
            for cell in col:
                if cell.row == 8:  # 8행이면
                    cell.fill = PatternFill(fgColor='E2EFDA', fill_type='solid')

        # 테두리 설정
        border_thin = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # 타이틀 영역
        for col in ws.iter_cols(min_row=2, min_col=2, max_row=3, max_col=3):
            for cell in col:
                cell.border = border_thin

        # context 영역
        for col in ws.iter_cols(min_row=8, min_col=2, max_row=43, max_col=6):
            for cell in col:
                cell.border = border_thin

        # 파일 저장하기
        print('스타일 적용 완료')


if __name__ == '__main__':
    wwp = WeeklyWorkPlan(start_date='2022-09-01', manager='김경록')
    wwp.save('주간업무계획표.xlsx')
