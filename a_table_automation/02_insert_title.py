from openpyxl import load_workbook, Workbook

# workbook 만들기
def create_excel_file():
    wb = Workbook()
    new_file = '주간업무계획표.xlsx'

    wb.save(new_file)
    print('엑셀파일 생성 완료')

def set_title():
    path = '주간업무계획표.xlsx'

    wb = load_workbook(path)
    wb.active.title = '주간업무계획표'

    # 현재 활성화 되어있는 워크시트 기본값 1번째
    ws = wb.active
    ws.cell(row=2, column=2, value='담당자')

    # ws.cell(2, 2, value='담당자')
    ws['C2'] = '김경록'
    ws['B3'] = '시작일'
    ws['C3'] = '2022-08-14'

    ws['B5'] = '주간업무계획표'
    ws['B6'] = '(2022-08-14~2022-08-20)'

    # 셀병합
    ws.merge_cells('B5:F5')
    ws.merge_cells('B6:F6')

    wb.save(path)
    print('타이틀 생성 완료')

