from openpyxl import load_workbook, Workbook


# workbook 만들기
def create_excel_file(filename):
    wb = Workbook()

    wb.save(filename)
    print('엑셀파일 생성 완료')


def set_title(filename):

    wb = load_workbook(filename)
    wb.active.title = '주간업무계획표'

    # 현재 활성화 되어있는 워크시트 기본값 1번째
    ws = wb.active
    ws.cell(row=2, column=2, value='담당자')

    # ws.cell(2, 2, value='담당자')
    ws['C2'] = '김경록'
    ws['B3'] = '시작일'
    ws['C3'] = '2022-08-14'

    ws['B5'] = '주간업무계획표'
    ws['B6'] = '(2022-08-14~2022-08-20)'

    # 셀병합
    ws.merge_cells('B5:F5')
    ws.merge_cells('B6:F6')

    wb.save(filename)
    print('타이틀 생성 완료')


def insert_context(filename):
    wb = load_workbook(filename)

    ws = wb.active

    cols_data = ['날짜', '요일', '시간', '일정', '비고']

    for col_idx in range(len(cols_data)):
        # print(col)
        ws.cell(row=8, column=2 + col_idx).value = cols_data[col_idx]

    week_date = ['8/14', '8/15', '8/16', '8/17', '8/18', '8/19', '8/20']

    k=9
    s=2
    for week in week_date:
        ws.cell(row=k, column=s).value = week
        k=k+1


    weekdays = ['Monday', 'Tuesday', 'Wendsday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    v=9
    m=3
    for day in weekdays:
        ws.cell(row=v, column=m).value = day
        v=v+1

    # 행 중간 삽입
    ws.insert_rows(10, 4)
    ws.insert_rows(15, 4)
    ws.insert_rows(20, 4)
    ws.insert_rows(25, 4)
    ws.insert_rows(30, 4)
    ws.insert_rows(35, 4)
    ws.insert_rows(40, 4)



    # 날짜, 요일, 비고 셀 병합

    ws.merge_cells('B9:B13')#날짜
    ws.merge_cells('C9:C13')#요일
    ws.merge_cells('F9:F13')#비고

    ws.merge_cells('B14:B18')#날짜
    ws.merge_cells('C14:C18')#요일
    ws.merge_cells('F14:F18')#비고

    ws.merge_cells('B19:B23')#날짜
    ws.merge_cells('C19:C23')#요일
    ws.merge_cells('F19:F23')#비고

    ws.merge_cells('B24:B28')#날짜
    ws.merge_cells('C24:C28')#요일
    ws.merge_cells('F24:F28')#비고

    ws.merge_cells('B29:B33')#날짜
    ws.merge_cells('C29:C33')#요일
    ws.merge_cells('F29:F33')#비고

    ws.merge_cells('B34:B38')#날짜
    ws.merge_cells('C34:C38')#요일
    ws.merge_cells('F34:F38')#비고

    ws.merge_cells('B39:B43')#날짜
    ws.merge_cells('C39:C43')#요일
    ws.merge_cells('F39:F43')#비고


    wb.save(filename)
    print('context 생성 완료')

if __name__ == '__main__':
    filename = '주간업무계획표.xlsx'
    create_excel_file(filename)
    set_title(filename)
    insert_context(filename)
