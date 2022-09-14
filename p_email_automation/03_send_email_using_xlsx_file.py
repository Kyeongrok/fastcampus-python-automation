import smtplib
from email.message import EmailMessage
from email.mime.text import MIMEText
from os import getenv

from openpyxl import load_workbook


class EmailSender:
    email_addr = None
    password = None
    smtp_server = None
    smtp_server_map = {
        'gmail.com': 'smtp.gmail.com',
        'naver.com': 'smtp.naver.com',
        'outlook.com': 'smtp-mail.outlook.com'
    }

    def __init__(self, email_addr, password):
        self.email_addr = email_addr
        self.password = password
        email_host = email_addr.split('@')[1]
        try:
            self.smtp_server = self.smtp_server_map[email_host]
        except Exception:
            raise ValueError(f'smtp_server_map에서 {email_host}를 찾을 수 없습니다')


    def send_email(self, msg, from_addr, to_addr, subject):
        """
        :param msg: 보낼 메세지
        :param from_addr: 보내는 사람
        :param to_addr: 받는 사람
        :return:
        """
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEText(msg)
            msg['From'] = from_addr
            msg['To'] = to_addr
            msg['Subject'] = subject
            print(msg.as_string())

            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print('이메일 전송이 완료 되었습니다.')

    def send_all_emails(self, filename):
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            print(row[0].value, row[1].value, row[2].value, row[3].value)
            self.send_email(from_addr=self.email_addr,
                      to_addr=row[0].value, subject=row[2].value, msg=row[3].value)



if __name__ == '__main__':
    email_addr = 'oceanfog@naver.com'
    password = getenv('MY_EMAIL_PASSWORD')
    es = EmailSender(email_addr, password)
    es.send_all_emails('이메일목록.xlsx')
