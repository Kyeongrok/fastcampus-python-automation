import smtplib
from email.message import EmailMessage
from email.mime.text import MIMEText
from email.utils import formataddr
from os import getenv

from openpyxl import load_workbook


class EmailSender:
    email_addr = None
    password = None
    smtp_server = None
    smtp_server_map = {
        'gmail.com': 'smtp.gmail.com',
        'naver.com': 'smtp.naver.com',
        'outlook.com': 'smtp-mail.outlook.com'
    }

    def __init__(self, email_addr, password):
        self.email_addr = email_addr
        self.password = password
        email_host = email_addr.split('@')[1]
        try:
            self.smtp_server = self.smtp_server_map[email_host]
        except Exception:
            raise ValueError(f'smtp_server_map에서 {email_host}를 찾을 수 없습니다')


    def send_email(self, msg, from_addr, to_addr, subject):
        """
        :param msg: 보낼 메세지
        :param from_addr: 보내는 사람
        :param to_addr: 받는 사람
        :return:
        """
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEText(msg)
            msg['From'] = formataddr(('김경록', self.email_addr))
            msg['To'] = to_addr
            msg['Subject'] = subject

            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print('이메일 전송이 완료 되었습니다.')

    def send_all_emails(self, filename):
        wb = load_workbook(filename)
        ws = wb.active

        tem1 = """
    안녕하세요. %받는분%님
    패스트몰 %담당자% 입니다.
    금일자로 주문건 접수되어 출고요청 전달드립니다.
    확인부탁드리겠습니다.
    항상 많은 도움주셔서 감사드립니다.
    %담당자% 드림
        """
        for row in ws.iter_rows(min_row=2):
            tem1 = tem1.replace('%받는분%', row[2].value)
            tem1 = tem1.replace('%담당자%', '김미미')
            print(row[0].value, row[1].value, row[2].value, tem1)
            self.send_email(from_addr=self.email_addr,
                      to_addr=row[0].value, subject=row[3].value, msg=tem1)



if __name__ == '__main__':
    email_addr = 'oceanfog@naver.com'
    password = getenv('MY_EMAIL_PASSWORD')
    es = EmailSender(email_addr, password)
    es.send_all_emails('이메일목록.xlsx')
