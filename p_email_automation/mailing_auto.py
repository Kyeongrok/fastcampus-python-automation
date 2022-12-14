import datetime
import smtplib
from pathlib import Path
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.encoders import encode_base64
from email.header import Header
from os import getenv


class SendEmail:
    # init은 인스턴스 객체 초기화
    def __init__(self, id, password):
        # 로그인 계정/pw
        self.id = id
        self.pw = password  # 앱비밀번호16자리
        smtp_server_map = {
            'gmail.com':'smtp.gmail.com',
            'naver.com':'smtp.naver.com',
        }
        self.smtp_server = smtp_server_map[id.split('@')[1]]


    def email_template(self, text, client_name, manager_name):
        with open('resources/' + text, encoding='UTF-8') as f:
            s = f.read()
            s = s.replace('업체명', client_name)
            s = s.replace('담당자명', manager_name)
        return s

    def make_email(self, row) -> MIMEMultipart:
        recipient = row[0].value
        # print(recipient)
        # # 수신자가 정해지지 않았을 경우 바로 넘어가게하기
        # if recipient is None:
        #     print("수신자 주소 오류")
        #     continue
        recipient2 = row[1].value
        title = row[2].value
        text = row[3].value
        attachment = row[4].value
        client_name = row[5].value
        manager_name = row[6].value
        print(recipient, recipient2, title, text, attachment, client_name, manager_name)
        msg = MIMEMultipart('alternative')
        to_list = [recipient]
        # 보내는 사람 / 받는 사람 / 참조 이메일 / 제목 입력
        msg["From"] = Header(self.id, 'utf-8')
        msg["To"] = ','.join(to_list)
        if recipient2 is not None:  # 참조이메일이 비어 있지 않으면 (!="")라고 할 경우 error 발생 AttributeError: 'NoneType' object has no attribute 'split'
            msg["Cc"] = Header(recipient2, 'utf-8')  # 슬래시 구분해서 input .split('/')
        # 제목 구성
        if title is None:  # 타이틀을 일일이 정해주지 않고 빈칸으로 내비두면 나오게 해주는 디폴트 문구 설정
            tx = "디폴트 제목"
            msg["Subject"] = Header(tx, 'utf-8')
        else:
            msg["Subject"] = Header(title, 'utf-8')
        # 본문 구성
        html_msg = self.email_template(text, client_name, manager_name)
        msg.attach(MIMEText(html_msg, 'html', 'utf-8'))
        # if text is None:  # 내용을 일일이 정해주지 않고 빈칸으로 내비두면 나오게 해주는 디폴트 문구 설정
        #     tx = "디폴트 내용"
        #     msg.set_content(tx)
        # else:
        #     msg.set_content(text)

        # 파일 첨부
        if attachment:
            part = MIMEBase('application', "octet-steam")
            filenm = Path(attachment).name
            with open('./data/' + attachment, 'rb') as f:
                part.set_payload(f.read())
                encode_base64(part)
                part.add_header('Content-Disposition', 'attachment; filename= "%s"' % filenm)
                msg.attach(part)
            #     msg.add_attachment(f.read(), maintype='application', subtype='octet-stream', filename=filenm)

        return msg

    def send_emails(self, excel_filename):
        # 엑셀파일에서 가져온 정보를 활용해 함수 반복 실행
        wb = load_workbook(excel_filename, data_only=True)
        ws = wb.active
        # read_email_list, for문 안에서 email 발송
        for row in ws.iter_rows(min_row=2):
            print(row)
            msg = self.make_email(row)
            self.send_email(msg, fr=self.id, to=row[0].value)

    def send_email(self, msg, fr, to):
        # 보내는 사람 로그인 및 smtp 서버로 발송
        # 메일 타입별 서버이름 변경 gmail, naver, outlook 아니면 오류 메세지 띄우게 설정
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            smtp.set_debuglevel(True)
            smtp.ehlo()
            smtp.starttls()
            smtp.login(self.id, self.pw)

            smtp.sendmail(to_addrs=to, from_addr=fr, msg=msg.as_string())
            smtp.quit()
            # 완료 메시지
            print("발송 성공")


# #인스턴스 생성
email = 'oceanfog1@gmail.com'
password = getenv('MY_EMAIL_PASSWORD')

es = SendEmail(email, password)  # 생성된 이메일리스트 따로 입력하지 않아도 자동입력, 계정 pw만 외부에서 입력받기
# #메소드 호출
es.send_emails('email_list.xlsx')
